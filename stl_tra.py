import sys
import os
from PyQt5.uic import loadUi
from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QDialog, QListWidget, QStackedWidget, QApplication, QWidget, QFileDialog, QTextEdit, QPushButton, QLabel
from PyQt5.QtCore import QDir
from stl import mesh
import math
import numpy
from matplotlib import pyplot
from mpl_toolkits import mplot3d
import time
import xlrd
import openpyxl
from openpyxl import load_workbook
from PyQt5.QtWidgets import *
from PyQt5 import QtCore, QtGui
from PyQt5.QtGui import *
from PyQt5.QtCore import *


class StlTranslator(QDialog):
    def __init__(self):
        super().__init__()
        loadUi("stl_tra.ui", self)
        
        
        self.loadstl.clicked.connect(self.explorestl)
        self.loadposition.clicked.connect(self.explorepos)
        self.generate.clicked.connect(self.generatestl)
        self.spinBox.setValue(1)
        self.loadstl_2.clicked.connect(self.explorestlgroup)
        self.clear.clicked.connect(self.clear_button)
        self.show()
		# showing all the widgets
        #self.modeCOG.clicked.connect(self.ModeCOG)
        #self.modeBB.clicked.connect(self.ModeBB)
        #self.modeCBB.clicked.connect(self.ModeCBB)
        
       # self.TextEditor_4.setPlainText(self.setting.currentText())
        
    def explorestl(self):
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.ExistingFiles)
        dialog.setFilter(QDir.Files)
        
        if dialog.exec_():
            file_name = dialog.selectedFiles()
            self.TextEditor.setPlainText(file_name[0])
            self.list_widget.clear()
    
    def explorestlgroup(self):
        
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.ExistingFiles)
        dialog.setFilter(QDir.Files)
        
        if dialog.exec_():
            file_name = dialog.selectedFiles()
            num_stl = len(file_name)
            self.TextEditor.setPlainText(file_name[0])
            for i in range(0,num_stl):
                
                item = QListWidgetItem(file_name[i])
                self.list_widget.addItem(item)
        
            self.list_widget.setDragDropMode(QAbstractItemView.InternalMove)                     
            self.TextEditor.clear()
           
    def explorepos(self):
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.AnyFile)
        dialog.setFilter(QDir.Files)
        
        if dialog.exec_():
            file_name = dialog.selectedFiles()
            self.TextEditor_2.setPlainText(file_name[0])
    
    def clear_button(self):
        self.TextEditor.clear()
        self.list_widget.clear()
        
            
            
             
     
                
    #def ModeBB(self):
        #self.TextEditor_4.setPlainText('Mode Bounding Box')
        
    #def ModeCOG(self):
        #self.TextEditor_4.setPlainText('Mode Center of Gravity')
        
    #def ModeCBB(self):
        #self.TextEditor_4.setPlainText('Mode Center of Bounding Box')    
                   
    def generatestl(self):
        #read positionfile
        if self.list_widget.count() ==0:
            
            i = self.spinBox.value()
            if self.setting_2.currentText() == '.txt':
                with open(self.TextEditor_2.toPlainText(), "r") as f:
                    content = f.readlines() 
                    f.close()
                    
                # xPos =list(range(1))
                
                # yPos =list(range(1))
                
                
                # #for i in range(1):
                xPos = content[(i-1)*7+3-1].rstrip('\n')
                yPos = content[(i-1)*7+5-1].rstrip('\n')
                zPos = content[(i-1)*7+7-1].rstrip('\n')
                
                self.TextEditor_3.setPlainText(str(i) + ':' + xPos + yPos + zPos)
                
                x = xPos[3:]
                y = yPos[3:]
                z = zPos[3:]
                
                x = float(x)
                y = float(y)
                z = float(z)

            elif self.setting_2.currentText() == '.xlsx':
                workbook = load_workbook(self.TextEditor_2.toPlainText(),read_only=True, data_only=True)
                worksheet = workbook['Positions Grundkörp (Copy paste' ]
                x = worksheet[chr(65+i)+str(40)].value
                y = worksheet[chr(65+i)+str(41)].value
                z = 0  
                
                x = float(x)
                y = float(y)
                z = float(z)
                
                self.TextEditor_3.setPlainText(str(i) + ':' + 'x='+ str(x) +'y=' + str(y) + 'z=' + str(z)) 
            
            your_mesh = mesh.Mesh.from_file(self.TextEditor.toPlainText())#mesh input
            
            
            volume, cog, inertia = your_mesh.get_mass_properties()
            # # # #print(cog)# current cog

            # # # #translate
            if self.setting.currentText() == 'Base on the Center of Gravity' :
                your_mesh.translate(-cog)
                your_mesh.translate([x,y,z])
                
            elif self.setting.currentText() == 'Base on the Bounding Box':
                minx = your_mesh.x.min()
                miny = your_mesh.y.min()
                minz = your_mesh.z.min()
                your_mesh.translate([-minx, -miny, -minz])
                your_mesh.translate([x,y,z])
            
            elif self.setting.currentText() == 'Base on the Center of Bounding Box':
                minx = your_mesh.x.min()
                maxx = your_mesh.x.max()
                miny = your_mesh.y.min()
                maxy = your_mesh.y.max()
                minz = your_mesh.z.min()
                maxz = your_mesh.z.max()
                your_mesh.translate([-(minx+maxx)/2, -(miny+maxy)/2, -(minz+maxz)/2])
                your_mesh.translate([x,y,z])
            
            t = time.strftime('%H%M%S', time.localtime()) 
            # # # #save
            your_mesh.save(self.TextEditor.toPlainText() +'_'+ str(i)+ '_' + 'results_'+ str(t) +'.stl')    
            
        elif self.list_widget.count() !=0:
            num_stl = self.list_widget.count()
            
            for i in range(0,num_stl):
                
                if self.setting_2.currentText() == '.txt':
                    with open(self.TextEditor_2.toPlainText(), "r") as f:
                        content = f.readlines() 
                        f.close()
                        
                    # xPos =list(range(1))
                    
                    # yPos =list(range(1))
                    
                    
                    # #for i in range(1):
                    xPos = content[(i)*7+3-1].rstrip('\n')
                    yPos = content[(i)*7+5-1].rstrip('\n')
                    zPos = content[(i)*7+7-1].rstrip('\n')
                    
                    self.TextEditor_3.append(str(i+1) + ':' + xPos + yPos + zPos)
                    
                    x = xPos[3:]
                    y = yPos[3:]
                    z = zPos[3:]
                    
                    x = float(x)
                    y = float(y)
                    z = float(z)           
            
                    
                elif self.setting_2.currentText() == '.xlsx':
                    workbook = load_workbook(self.TextEditor_2.toPlainText(),read_only=True, data_only=True)
                    worksheet = workbook['Positions Grundkörp (Copy paste' ]
                    x = worksheet[chr(66+i)+str(40)].value
                    y = worksheet[chr(66+i)+str(41)].value
                    z = 0  
                
                    x = float(x)
                    y = float(y)
                    z = float(z)
                    
                    self.TextEditor_3.append(str(i+1) + ':' + 'x='+ str(x) +'y=' + str(y) + 'z=' + str(z))
                
                    
                your_mesh = mesh.Mesh.from_file(self.list_widget.item(i).text())#mesh input
            
            
                volume, cog, inertia = your_mesh.get_mass_properties()
                
                
                your_mesh.translate(-cog)
                your_mesh.translate([x,y,z])
                # # # #translate
                if self.setting.currentText() == 'Base on the Center of Gravity' :
                    your_mesh.translate(-cog)
                    your_mesh.translate([x,y,z])
                    
                elif self.setting.currentText() == 'Base on the Bounding Box':
                    minx = your_mesh.x.min()
                    miny = your_mesh.y.min()
                    minz = your_mesh.z.min()
                    your_mesh.translate([-minx, -miny, -minz])
                    your_mesh.translate([x,y,z])
                
                elif self.setting.currentText() == 'Base on the Center of Bounding Box':
                    minx = your_mesh.x.min()
                    maxx = your_mesh.x.max()
                    miny = your_mesh.y.min()
                    maxy = your_mesh.y.max()
                    minz = your_mesh.z.min()
                    maxz = your_mesh.z.max()
                    your_mesh.translate([-(minx+maxx)/2, -(miny+maxy)/2, -(minz+maxz)/2])
                    your_mesh.translate([x,y,z])
                
                else:
                    self.qWarning('no Mode')
                
                
                t = time.strftime('%H%M%S', time.localtime()) 
                # # # #save
                your_mesh.save(self.list_widget.item(i).text() +'_' + str(i+1)+ '_' + 'results_'+ str(t) +'.stl') 
                
        


            
        

        
        
# main
app = QApplication(sys.argv)
stltra = StlTranslator()
widget = QStackedWidget()
widget.addWidget(stltra)
widget.setFixedHeight(800)
widget.setFixedWidth(1200)
widget.show()
sys.exit(app.exec_())
